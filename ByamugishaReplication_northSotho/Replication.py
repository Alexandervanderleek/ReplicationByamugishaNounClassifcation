import fasttext
import openpyxl
from utility import get_semantic_noun_neigbours, get_unique_prefixes, is_unique, most_frequent
from gensim.models.fasttext import load_facebook_model
from gensim.similarities.annoy import AnnoyIndexer

'''
Joans Model for northern Sotho sepedi
'''

workbook = openpyxl.load_workbook(f'./DataFiles/NorthernSothoSingle.xlsx')
worksheet = workbook.active

# Get all the class labels from the second column
nouns = [row[0].lower() for row in worksheet.iter_rows(min_row=1, values_only=True)]

workbook = openpyxl.load_workbook(f'./DataFiles/NorthernSothoSingle20%TestSet.xlsx')
worksheet = workbook.active

# Get all the class labels from the second column
test_nouns_and_classes = [(row[0].lower(),row[1]) for row in worksheet.iter_rows(min_row=1, values_only=True)]

workbook.close()

prefix_dictionary = [
    ('1',"mo"),
    ('1a',""),
    ('2',"ba"),
    ('2b',"bo"),
    ('3',"mo"),
    ('4',"me"),
    ('5',"le"),
    ('6',"ma"),
    ('7',"se"),
    ('8',"di"),
    ('9',"n"),
    ('9',"m"),
    ('9',""),
    ('10',"din"),
    ('10',"dim"),
    ('10',"di"),
    ('14',"bo"),
    ('16',"fa"),
    ('17',"go"),
    ('18',"mo"),
    ('n',"m"),
    ('n',"n"),
    ('n',""),
    ('24',"ga"),
]

#load the semantic model
semantic_model = load_facebook_model('./ModelsAndVectors/nSotho_fasttext_embeddings.bin')

#annoy index
annoyIndex = AnnoyIndexer(model=semantic_model, num_trees=10)

#load classification model
classifier_model = fasttext.load_model('./ModelsAndVectors/nSotho_fasttext_classifer.bin')

#get unique prefixes from list
unique_prefixes_classes = get_unique_prefixes(prefixDict=prefix_dictionary)

correct=0
guesses=0


for query_word, correct_class in test_nouns_and_classes:
    guesses+=1
    print(str(guesses), end='\r')

    #check if unique prefix exists
    value = is_unique(unique_prefixes_classes, query=query_word)

    if(value):
        if correct_class==value:
            correct+=1
        else:
            continue

    else:
        # get semantic noun neighbours for query
        semantic_similar_words = get_semantic_noun_neigbours(nouns=nouns,semantic_model=semantic_model,indexer=annoyIndex, query_word=query_word,topn=20)

        predicted_labels = []
        
        for word in semantic_similar_words:
            #make a prediction using classifer
            prediction = classifier_model.predict(word, k=20)

            labels, arr = prediction

            labels = list(labels)
            
            #do filtering for final noun class and subject concord

            noun_labels = filter((lambda x: "SC" not in x),labels)
            concord_labels = filter((lambda x: "SC" in x),labels)
            

            final_noun_class = list(noun_labels)[0].replace("__label__","")
            final_concord_class = list(concord_labels)[0].replace("__label__SC","")

            if final_noun_class == final_concord_class:
                predicted_labels.append(final_concord_class)
            
        #make a final prediction based on consensus
        predicted_label = most_frequent(predicted_labels)
        if correct_class == predicted_label:
            correct += 1

print(guesses)
print(correct)
print(str(correct/guesses))
